#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
02_build_env_manual_label_file.py
==================================
BƯỚC 2 — Tạo file template để người nghiên cứu gán nhãn env_human.

Đề tài: Đánh giá tác động chính sách môi trường — Pipeline 2 tầng
Tác giả: Đinh Công Khang — MI3380 Đồ án 1

Vị trí trong pipeline:
    Input:  data/raw/Luat_bao_ve_moi_truong_2020.json
    Output: data/interim/env_manual_label_template.xlsx
    Tiếp theo: [HUMAN gán env_human] → Script 03

Lý do cần bước này:
    - Người nghiên cứu cần file Excel có cột trống để điền nhãn
    - File phải có đủ thông tin bản ghi (raw_text, legal_citation, actor, ...)
      để người gán có thể đưa ra quyết định mà không cần tra cứu thêm
    - Script tự động highlight các bản ghi có keyword môi trường
      (dùng env_filter_keywords.yaml) để hỗ trợ gán nhanh hơn
    - KHÔNG tự động gán nhãn — chỉ tạo template trống

Hướng dẫn sau khi chạy:
    1. Mở data/interim/env_manual_label_template.xlsx
    2. Đọc docs/environmental_filter_guideline.md TRƯỚC khi gán
    3. Điền vào cột 'env_human': 1 (có tác động MTs) hoặc 0 (không có)
    4. Điền 'env_human_reason': lý do + domain nếu env_human=1
    5. Điền 'env_needs_review': True nếu còn mơ hồ
    6. Lưu file thành data/interim/env_human_labeled_dataset.xlsx

Sử dụng:
    py src/02_build_env_manual_label_file.py
    py src/02_build_env_manual_label_file.py --input data/raw/my_law.json
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml

# ─── Project root ──────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# ─── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except AttributeError:
        pass


# ==============================================================================
# Load config
# ==============================================================================

def load_config():
    config_path = PROJECT_ROOT / "config" / "project_config.yaml"
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_keywords():
    kw_path = PROJECT_ROOT / "config" / "env_filter_keywords.yaml"
    with open(kw_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    
    # Gộp tất cả keyword từ các domain thành 1 set (lowercase)
    all_keywords = set()
    for domain, info in cfg.get("environment_domains", {}).items():
        for kw in info.get("keywords", []):
            all_keywords.add(kw.lower())
    return all_keywords


def load_json_data(filepath: Path) -> list:
    import json
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    elif isinstance(data, dict) and "records" in data:
        return data["records"]
    return [data]


# ==============================================================================
# Xử lý chính
# ==============================================================================

def detect_keywords(text: str, keywords: set) -> str:
    """Tìm các keyword môi trường trong text, trả về chuỗi các keyword tìm thấy."""
    if not text:
        return ""
    text_lower = text.lower()
    found = [kw for kw in sorted(keywords) if kw in text_lower]
    return " | ".join(found[:10])  # Giới hạn 10 keyword để không quá dài


def build_template(records: list, keywords: set, cfg: dict) -> pd.DataFrame:
    """
    Xây dựng DataFrame template từ danh sách bản ghi.

    Cột đầu ra:
        - Các cột thông tin bản ghi (source_id, legal_citation, raw_text, ...)
        - Cột hỗ trợ: keyword_found (tự động highlight)
        - Cột cần điền (highlight vàng): env_human, env_domain_human,
          env_human_reason, env_needs_review
        - Cột review (highlight xanh nhạt): review_note

    Căn cứ: docs/env_filter_guideline_cap_nhat.md — Mục 2 (trường dữ liệu)
    và Mục 6 (danh mục domain môi trường).
    """
    rows = []
    for rec in records:
        raw_text = str(rec.get("raw_text", ""))
        kw_found = detect_keywords(raw_text, keywords)

        row = {
            # ── Thông tin bản ghi ──
            "source_id":          rec.get("source_id", ""),
            "legal_citation":     rec.get("legal_citation", ""),
            "raw_text":           raw_text,
            "actor":              rec.get("actor", ""),
            "legal_signal":       rec.get("legal_signal", rec.get("tin_hieu_tac_dong", "")),
            "domain":             rec.get("domain", ""),
            "quantitative_value": rec.get("quantitative_value", rec.get("gia_tri_dinh_luong", "")),
            "condition":          rec.get("condition", rec.get("dieu_kien_ap_dung", "")),

            # ── Hỗ trợ gán nhãn (tự động) ──
            "keyword_found":      kw_found,           # Keyword MT tìm thấy trong raw_text
            "has_keyword":        1 if kw_found else 0,  # Flag: 1 = có keyword

            # ── Cần điền thủ công (highlight vàng) ──
            "env_human":          "",   # 1 hoặc 0 (xem Mục 4–5 guideline)
            "env_domain_human":   "",   # Domain MT: water; waste; air_noise_radiation; ...
                                        # (xem Mục 6 guideline để chọn domain chuẩn)
            "env_human_reason":   "",   # Lý do ngắn gọn (xem Mục 8–9 guideline)
            "env_needs_review":   "",   # TRUE nếu còn mơ hồ (xem Mục 3.4 guideline)

            # ── Chỉ điền khi review sau adjudication (highlight xanh nhạt) ──
            "review_note":        "",   # Ghi chú khi sửa nhãn sau khi đối chiếu LLM
        }
        rows.append(row)
    
    return pd.DataFrame(rows)


def save_template(df: pd.DataFrame, output_path: Path) -> None:
    """Lưu template ra file Excel với định dạng dễ đọc."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="env_label_template", index=False)
        
        # Định dạng cột
        ws = writer.sheets["env_label_template"]
        
        # Đặt độ rộng cột
        # Thứ tự cột: A–J thông tin bản ghi + hỗ trợ,
        #             K–N cần điền thủ công, O review_note
        col_widths = {
            "A": 20,  # source_id
            "B": 30,  # legal_citation
            "C": 60,  # raw_text
            "D": 25,  # actor
            "E": 25,  # legal_signal
            "F": 20,  # domain
            "G": 15,  # quantitative_value
            "H": 25,  # condition
            "I": 40,  # keyword_found
            "J": 12,  # has_keyword
            "K": 12,  # env_human          ← cần điền
            "L": 40,  # env_domain_human   ← cần điền (THÊM MỚI)
            "M": 45,  # env_human_reason   ← cần điền
            "N": 15,  # env_needs_review   ← cần điền
            "O": 45,  # review_note        ← chỉ điền khi review (THÊM MỚI)
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Wrap text cho cột raw_text và keyword_found
        from openpyxl.styles import Alignment, PatternFill, Font
        wrap_align = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            row[2].alignment = wrap_align   # raw_text (cột C, index 2)
            row[8].alignment = wrap_align   # keyword_found (cột I, index 8)
            row[12].alignment = wrap_align  # env_human_reason (cột M, index 12)

        # Highlight header các cột cần điền thủ công — màu vàng (K, L, M, N)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        for col_idx in [11, 12, 13, 14]:  # K, L, M, N (1-indexed)
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = yellow_fill
            cell.font = bold_font

        # Highlight header cột review_note — màu xanh nhạt (O)
        # Phân biệt với cột cần điền ngay: review_note chỉ điền sau adjudication
        blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        review_cell = ws.cell(row=1, column=15)  # cột O
        review_cell.fill = blue_fill
        review_cell.font = bold_font
    
    logger.info(f"  ✅ Đã lưu template: {output_path} ({len(df):,} dòng)")


# ==============================================================================
# Main
# ==============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Tạo file template để người nghiên cứu gán nhãn env_human"
    )
    parser.add_argument(
        "--input",
        type=str,
        default=None,
        help="Đường dẫn đến file JSON đầu vào (mặc định: từ project_config.yaml)"
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Đường dẫn file output .xlsx (mặc định: data/interim/env_manual_label_template.xlsx)"
    )
    return parser.parse_args()


def main():
    args = parse_args()
    cfg = load_config()
    
    logger.info("=" * 60)
    logger.info("  BƯỚC 2 — Tạo file template gán nhãn env_human")
    logger.info("=" * 60)
    
    # ── Đường dẫn input ──
    if args.input:
        input_path = Path(args.input)
        if not input_path.is_absolute():
            input_path = PROJECT_ROOT / input_path
    else:
        raw_dir = PROJECT_ROOT / cfg["paths"]["raw_data"]
        input_path = raw_dir / cfg["input"]["raw_json_file"]
    
    # ── Đường dẫn output ──
    if args.output:
        output_path = Path(args.output)
        if not output_path.is_absolute():
            output_path = PROJECT_ROOT / output_path
    else:
        interim_dir = PROJECT_ROOT / cfg["paths"]["interim_data"]
        output_path = interim_dir / cfg["interim_files"]["env_manual_label_template"]
    
    logger.info(f"[Bước 2.1] Đọc dữ liệu từ: {input_path}")
    if not input_path.exists():
        logger.error(f"File không tồn tại: {input_path}")
        sys.exit(1)
    
    records = load_json_data(input_path)
    logger.info(f"  → Tổng số bản ghi: {len(records):,}")
    
    if len(records) != 581:
        logger.warning(f"  ⚠️  Expected N=581, thực tế N={len(records)}. Kiểm tra lại dữ liệu đầu vào!")
    
    logger.info("[Bước 2.2] Load từ khóa môi trường từ env_filter_keywords.yaml")
    keywords = load_keywords()
    logger.info(f"  → Tổng số keyword: {len(keywords)}")
    
    logger.info("[Bước 2.3] Xây dựng template...")
    df = build_template(records, keywords, cfg)
    
    # Thống kê keyword
    n_has_keyword = int(df["has_keyword"].sum())
    logger.info(f"  → Bản ghi có keyword môi trường: {n_has_keyword}/{len(df)} ({n_has_keyword/len(df)*100:.1f}%)")
    logger.info(f"  → Bản ghi không có keyword:      {len(df)-n_has_keyword}/{len(df)}")
    logger.info("  ℹ️  Keyword chỉ là hỗ trợ — không quyết định env_human!")
    
    logger.info(f"[Bước 2.4] Lưu template ra: {output_path}")
    save_template(df, output_path)
    
    logger.info("")
    logger.info("=" * 60)
    logger.info("  HOÀN THÀNH — Bước tiếp theo:")
    logger.info(f"  1. Đọc docs/environmental_filter_guideline.md")
    logger.info(f"  2. Mở {output_path.name} và điền env_human (1/0)")
    logger.info(f"  3. Lưu lại thành: env_human_labeled_dataset.xlsx")
    logger.info(f"  4. Chạy: py src/03_llm_filter_environmental_impact.py")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
