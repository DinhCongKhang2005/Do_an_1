#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
07_build_scoring_input.py
==========================
BƯỚC 8 — Tạo bảng nhập biến tính điểm M_i, S_i, D_i, R_i, W_i.

Đề tài: Đánh giá tác động chính sách môi trường — Phiên bản 5 nhãn
Tác giả: Đinh Công Khang — MI3380 Đồ án 1

Pipeline:
  Input:  data/processed/final_labeled_dataset.xlsx  (có cột final_label)
  Output: data/interim/scoring_input.xlsx

Sau khi chạy script này, người nghiên cứu cần điền thủ công:
  M_i : Magnitude — cường độ tác động  (1–5)
  S_i : Scope     — phạm vi tác động   (1–5)
  D_i : Duration  — thời gian tác động (1–5)
  R_i : Risk/Reversibility              (1–5)
  W_i : Object/domain weight            (mặc định 1.0)

Xem docs/variable_definition.md và docs/scoring_rule.md để biết cách chấm điểm.

Sử dụng:
  python src/07_build_scoring_input.py
  python src/07_build_scoring_input.py --help
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml

PROJECT_ROOT = Path(__file__).resolve().parent.parent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

VALID_LABELS_5 = [
    "BENEFIT_QUANTITATIVE",
    "BENEFIT_QUALITATIVE",
    "COST_QUANTITATIVE",
    "COST_QUALITATIVE",
    "CONSTRAINT",
]


def load_config(config_path: Path) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def build_scoring_template(df: pd.DataFrame, default_weight: float = 1.0) -> pd.DataFrame:
    """
    Xây dựng bảng nhập biến tính điểm từ final_labeled_dataset.
    Chỉ lấy các record có final_label hợp lệ.
    """
    # Lọc record có final_label hợp lệ
    mask = df["final_label"].isin(VALID_LABELS_5) if "final_label" in df.columns else pd.Series([True] * len(df))
    df_valid = df[mask].copy()

    if len(df_valid) < len(df):
        skipped = len(df) - len(df_valid)
        logger.warning(f"Bỏ qua {skipped} record có final_label không hợp lệ (UNKNOWN/NaN)")

    # Chọn cột context
    context_cols = [
        "source_id", "legal_citation", "raw_text",
        "final_label", "final_label_source",
        "human_label", "llm_label", "label_match",
        "chu_the", "domain", "gia_tri_dinh_luong",
    ]
    existing = [c for c in context_cols if c in df_valid.columns]
    df_template = df_valid[existing].copy()

    # Thêm cột biến tính điểm (để trống — người nghiên cứu điền)
    df_template["M_i"] = None   # Magnitude
    df_template["S_i"] = None   # Scope
    df_template["D_i"] = None   # Duration
    df_template["R_i"] = None   # Risk/Reversibility
    df_template["W_i"] = default_weight  # Object weight (điền sẵn mặc định)

    return df_template


def export_scoring_excel(df: pd.DataFrame, output_path: Path):
    """Xuất bảng nhập điểm với định dạng hỗ trợ nhập liệu."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="scoring_input")
        ws = writer.sheets["scoring_input"]

        from openpyxl.utils import get_column_letter
        col_names = [cell.value for cell in ws[1]]

        col_widths = {
            "source_id": 22,
            "legal_citation": 22,
            "raw_text": 80,
            "final_label": 25,
            "final_label_source": 15,
            "human_label": 25,
            "llm_label": 25,
            "label_match": 12,
            "chu_the": 25,
            "domain": 20,
            "gia_tri_dinh_luong": 20,
            "M_i": 8,
            "S_i": 8,
            "D_i": 8,
            "R_i": 8,
            "W_i": 8,
        }
        for col_idx, col_name in enumerate(col_names, start=1):
            col_letter = get_column_letter(col_idx)
            width = col_widths.get(col_name, 15)
            ws.column_dimensions[col_letter].width = width

        # Đóng băng dòng tiêu đề
        ws.freeze_panes = "A2"

        # Sheet hướng dẫn chấm điểm
        ws_guide = writer.book.create_sheet("Hướng dẫn chấm điểm")
        guide_rows = [
            ["=== HƯỚNG DẪN CHẤM ĐIỂM BIẾN TÍNH ĐIỂM ==="],
            ["Thang điểm Likert: 1 (thấp nhất) → 5 (cao nhất)"],
            [""],
            ["Biến", "Tên đầy đủ", "Gợi ý chấm điểm"],
            ["M_i", "Magnitude (Cường độ tác động)",
             "1=không đáng kể, 3=trung bình, 5=rất lớn"],
            ["S_i", "Scope (Phạm vi tác động)",
             "1=cá nhân/hẹp, 3=ngành/vùng, 5=toàn quốc/toàn cầu"],
            ["D_i", "Duration (Thời gian tác động)",
             "1=tạm thời(<1 năm), 3=trung hạn, 5=vĩnh viễn/dài hạn"],
            ["R_i", "Risk/Reversibility",
             "1=có thể đảo ngược dễ, 3=khó đảo ngược, 5=không thể đảo ngược"],
            ["W_i", "Object/Domain Weight",
             "Mặc định 1.0. Có thể điều chỉnh nếu một domain quan trọng hơn"],
            [""],
            ["=== CÔNG THỨC TÍNH ==="],
            ["C_i = 0.25*M_i + 0.25*S_i + 0.25*D_i + 0.25*R_i"],
            ["C_i_norm = (C_i - 1) / 4   → C_i_norm ∈ [0,1]"],
            ["ImpactScore_i = direction_i * W_i * C_i_norm"],
            [""],
            ["direction_i = +1 cho BENEFIT_QUANTITATIVE, BENEFIT_QUALITATIVE"],
            ["direction_i = -1 cho COST_QUANTITATIVE, COST_QUALITATIVE, CONSTRAINT"],
            [""],
            ["Lưu file và chạy: python src/08_calculate_impact_score.py"],
        ]
        for row in guide_rows:
            ws_guide.append(row)
        ws_guide.column_dimensions["A"].width = 15
        ws_guide.column_dimensions["B"].width = 30
        ws_guide.column_dimensions["C"].width = 55

    logger.info(f"Đã xuất scoring input → {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Bước 8: Tạo bảng nhập biến M_i, S_i, D_i, R_i, W_i"
    )
    parser.add_argument("--config", type=Path,
                        default=PROJECT_ROOT / "config" / "project_config.yaml")
    parser.add_argument("--input", type=Path, default=None,
                        help="File final_labeled_dataset.xlsx (ghi đè config)")
    parser.add_argument("--output", type=Path, default=None,
                        help="File scoring_input.xlsx (ghi đè config)")
    args = parser.parse_args()

    logger.info("=" * 65)
    logger.info("  BƯỚC 8: TẠO BẢNG NHẬP BIẾN TÍNH ĐIỂM")
    logger.info(f"  Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 65)

    config = load_config(args.config)
    processed_dir = PROJECT_ROOT / config["paths"]["processed_data"]
    interim_dir = PROJECT_ROOT / config["paths"]["interim_data"]
    default_weight = config["scoring"].get("default_object_weight", 1.0)

    input_path = args.input or (processed_dir / config["processed_files"]["final_labeled_dataset"])
    output_path = args.output or (interim_dir / config["interim_files"]["scoring_input"])

    if not input_path.exists():
        logger.error(f"Không tìm thấy file: {input_path}")
        logger.error("Chạy trước: python src/06_compare_llm_vs_human.py")
        sys.exit(1)

    df = pd.read_excel(input_path)
    logger.info(f"Đọc được {len(df)} bản ghi từ: {input_path}")

    df_template = build_scoring_template(df, default_weight)
    export_scoring_excel(df_template, output_path)

    logger.info("")
    logger.info("✅ BƯỚC 8 HOÀN TẤT")
    logger.info(f"   Số bản ghi cần chấm điểm: {len(df_template)}")
    logger.info(f"   File template: {output_path}")
    logger.info("")
    logger.info("   === BƯỚC THỦ CÔNG TIẾP THEO ===")
    logger.info("   1. Mở file: data/interim/scoring_input.xlsx")
    logger.info("   2. Xem sheet 'Hướng dẫn chấm điểm'")
    logger.info("   3. Điền M_i, S_i, D_i, R_i (thang 1–5)")
    logger.info("   4. Giữ W_i = 1.0 hoặc điều chỉnh theo domain")
    logger.info("   5. Lưu file lại (cùng tên)")
    logger.info("")
    logger.info("   Sau khi hoàn thành:")
    logger.info("   python src/08_calculate_impact_score.py")


if __name__ == "__main__":
    main()
