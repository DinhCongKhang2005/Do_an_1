#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
06_compare_llm_vs_human.py
===========================
BƯỚC 7 — So sánh nhãn LLM với nhãn người và tính đầy đủ 8 metrics (5 nhãn).

Đề tài: Đánh giá tác động chính sách môi trường — Phiên bản 5 nhãn
Tác giả: Đinh Công Khang — MI3380 Đồ án 1

Pipeline:
  Input:  data/processed/llm_labeled_dataset.xlsx   (có cột llm_label)
          data/interim/human_labeled_dataset.xlsx    (có cột human_label)
  Output:
    - data/reports/classification_metrics.xlsx   (8 metrics đầy đủ)
    - data/reports/error_analysis.xlsx           (các record LLM sai)
    - data/processed/final_labeled_dataset.xlsx  (có final_label + label_match)

8 Metrics bắt buộc (theo README_3_updated Mục 7):
  1. Confusion Matrix (5x5)
  2. Accuracy
  3. Precision per-class
  4. Recall per-class
  5. F1-score per-class
  6. Macro-F1
  7. Support per-class
  8. Human Correction Rate

Sử dụng:
  python src/06_compare_llm_vs_human.py
  python src/06_compare_llm_vs_human.py --help
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import yaml
from sklearn.metrics import (
    accuracy_score,
    confusion_matrix,
    f1_score,
    precision_score,
    recall_score,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# Thứ tự nhãn trong Confusion Matrix 5x5
LABELS_5 = [
    "BENEFIT_QUANTITATIVE",
    "BENEFIT_QUALITATIVE",
    "COST_QUANTITATIVE",
    "COST_QUALITATIVE",
    "CONSTRAINT",
]
# Nhãn viết tắt cho hiển thị
LABEL_ABBR = {
    "BENEFIT_QUANTITATIVE": "BQ",
    "BENEFIT_QUALITATIVE":  "BQL",
    "COST_QUANTITATIVE":    "CQ",
    "COST_QUALITATIVE":     "CQL",
    "CONSTRAINT":           "CON",
}


def load_config(config_path: Path) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def merge_datasets(df_llm: pd.DataFrame, df_human: pd.DataFrame) -> pd.DataFrame:
    """
    Ghép nhãn LLM và nhãn người theo source_id.
    Nếu file LLM đã có human_label (do đọc từ human_labeled_dataset),
    loại bỏ cột cũ trước khi merge để tránh _x/_y.
    """
    if "source_id" not in df_llm.columns or "source_id" not in df_human.columns:
        logger.warning("Không tìm thấy cột 'source_id' — ghép theo chỉ số dòng")
        df_merged = df_llm.copy()
        if "human_label" in df_human.columns:
            min_len = min(len(df_llm), len(df_human))
            df_merged = df_merged.iloc[:min_len].copy()
            df_merged["human_label"] = df_human["human_label"].values[:min_len]
        return df_merged

    if "human_label" not in df_human.columns:
        logger.warning("File nhãn người không có cột 'human_label' — bỏ qua so sánh")
        return df_llm.copy()

    # Loại bỏ cột human_* cũ khỏi df_llm (nếu có) để tránh trùng
    stale = [c for c in ("human_label", "human_reason", "review_note", "needs_second_review")
             if c in df_llm.columns]
    df_llm_base = df_llm.drop(columns=stale) if stale else df_llm

    human_cols = ["source_id", "human_label"]
    for opt in ("human_reason", "review_note", "needs_second_review"):
        if opt in df_human.columns:
            human_cols.append(opt)

    df_human_sub = df_human[[c for c in human_cols if c in df_human.columns]]
    df_merged = df_llm_base.merge(df_human_sub, on="source_id", how="left")
    n_matched = df_merged["human_label"].notna().sum()
    logger.info(f"Ghép nhãn người: {n_matched}/{len(df_merged)} bản ghi có human_label")
    return df_merged


def compute_all_metrics(y_true: np.ndarray, y_pred: np.ndarray) -> dict:
    """
    Tính đầy đủ 8 metrics theo đặc tả README Mục 7.
    Chỉ xét các record có cả hai nhãn hợp lệ (trong LABELS_5).
    """
    valid_set = set(LABELS_5)
    mask = (
        pd.Series(y_true).isin(valid_set).values
        & pd.Series(y_pred).isin(valid_set).values
    )
    y_true_f = np.array(y_true)[mask]
    y_pred_f = np.array(y_pred)[mask]

    M_eval = len(y_true_f)
    if M_eval == 0:
        logger.error("Không có record nào có cả human_label và llm_label hợp lệ!")
        return {}

    # ── 1. Accuracy ──────────────────────────────────────────────────────────
    accuracy = accuracy_score(y_true_f, y_pred_f)

    # ── 2-5. Precision, Recall, F1 per-class ─────────────────────────────────
    per_class = {}
    for lbl in LABELS_5:
        tp = int(np.sum((y_true_f == lbl) & (y_pred_f == lbl)))
        fp = int(np.sum((y_true_f != lbl) & (y_pred_f == lbl)))
        fn = int(np.sum((y_true_f == lbl) & (y_pred_f != lbl)))
        support_k = int(np.sum(y_true_f == lbl))

        prec = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        rec = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1 = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0.0

        per_class[lbl] = {
            "TP": tp, "FP": fp, "FN": fn,
            "Precision": round(prec, 4),
            "Recall": round(rec, 4),
            "F1": round(f1, 4),
            "Support": support_k,
        }

    # ── 6. Macro-F1 ───────────────────────────────────────────────────────────
    macro_f1 = round(sum(v["F1"] for v in per_class.values()) / 5, 4)

    # ── 7. Support (đã tính trong per_class) ──────────────────────────────────

    # ── 8. Human Correction Rate ──────────────────────────────────────────────
    human_correction_rate = round(1.0 - accuracy, 4)

    # ── Confusion Matrix 5x5 ──────────────────────────────────────────────────
    cm = confusion_matrix(y_true_f, y_pred_f, labels=LABELS_5)

    # ── Label Match (per record) ──────────────────────────────────────────────
    label_match = (y_true_f == y_pred_f).tolist()

    return {
        "M_eval": M_eval,
        "M_total": len(y_true),
        "accuracy": round(float(accuracy), 4),
        "macro_f1": macro_f1,
        "human_correction_rate": human_correction_rate,
        "per_class": per_class,
        "confusion_matrix": cm.tolist(),
        "y_true_filtered": y_true_f.tolist(),
        "y_pred_filtered": y_pred_f.tolist(),
        "label_match": label_match,
        "mask": mask.tolist(),
    }


def print_metrics_summary(metrics: dict):
    """In tóm tắt metrics ra console."""
    logger.info(f"\n{'='*65}")
    logger.info("  KẾT QUẢ SO SÁNH LLM vs HUMAN (5 NHÃN)")
    logger.info(f"{'='*65}")
    logger.info(f"  Số bản ghi đánh giá (M_eval) : {metrics['M_eval']}/{metrics['M_total']}")
    logger.info(f"  Accuracy                      : {metrics['accuracy']:.4f}")
    logger.info(f"  Macro-F1                      : {metrics['macro_f1']:.4f}")
    logger.info(f"  Human Correction Rate         : {metrics['human_correction_rate']:.4f}")
    logger.info("")
    logger.info("  Theo từng nhãn (Precision / Recall / F1 / Support):")
    for lbl, v in metrics["per_class"].items():
        abbr = LABEL_ABBR.get(lbl, lbl[:5])
        logger.info(
            f"    {lbl:<25}: P={v['Precision']:.3f}  "
            f"R={v['Recall']:.3f}  F1={v['F1']:.3f}  n={v['Support']}"
        )
    logger.info("")
    logger.info("  Confusion Matrix 5x5 (hàng=Human, cột=LLM):")
    header = "  " + " " * 12 + "  ".join(f"{LABEL_ABBR[l]:>4}" for l in LABELS_5)
    logger.info(header)
    cm = np.array(metrics["confusion_matrix"])
    for i, lbl in enumerate(LABELS_5):
        row_str = "  ".join(f"{v:>4}" for v in cm[i])
        logger.info(f"    {LABEL_ABBR[lbl]:<6}: {row_str}")
    logger.info(f"{'='*65}\n")


def export_metrics_excel(metrics: dict, output_path: Path):
    """Xuất 8 metrics ra Excel 3 sheet."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Sheet 1: Tổng hợp
    summary_rows = [
        {"Metric": "Số bản ghi đánh giá (M_eval)", "Giá trị": metrics["M_eval"]},
        {"Metric": "Số bản ghi tổng (M_total)", "Giá trị": metrics["M_total"]},
        {"Metric": "Accuracy", "Giá trị": metrics["accuracy"]},
        {"Metric": "Macro-F1", "Giá trị": metrics["macro_f1"]},
        {"Metric": "Human Correction Rate (= 1 - Accuracy)", "Giá trị": metrics["human_correction_rate"]},
    ]
    df_summary = pd.DataFrame(summary_rows)

    # Sheet 2: Per-class metrics
    per_class_rows = []
    for lbl in LABELS_5:
        v = metrics["per_class"][lbl]
        per_class_rows.append({
            "Nhãn": lbl,
            "Precision": v["Precision"],
            "Recall": v["Recall"],
            "F1-score": v["F1"],
            "Support (n)": v["Support"],
            "TP": v["TP"],
            "FP": v["FP"],
            "FN": v["FN"],
        })
    df_per_class = pd.DataFrame(per_class_rows)

    # Sheet 3: Confusion Matrix 5x5
    cm = np.array(metrics["confusion_matrix"])
    col_labels = [f"Pred_{LABEL_ABBR[l]}" for l in LABELS_5]
    row_labels = [f"True_{LABEL_ABBR[l]}" for l in LABELS_5]
    df_cm = pd.DataFrame(cm, index=row_labels, columns=col_labels)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Tổng hợp")
        df_per_class.to_excel(writer, index=False, sheet_name="Chi tiết theo nhãn")
        df_cm.to_excel(writer, sheet_name="Confusion Matrix 5x5")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 18

    logger.info(f"Đã xuất metrics → {output_path}")


def export_error_analysis(df_merged: pd.DataFrame, output_path: Path):
    """
    Xuất các record có label_match = false ra file Error Analysis.
    Bao gồm phân tích theo các cặp lỗi ưu tiên.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    mask_valid = (
        df_merged["human_label"].isin(LABELS_5)
        & df_merged["llm_label"].isin(LABELS_5)
    )
    df_eval = df_merged[mask_valid].copy()
    df_errors = df_eval[df_eval["human_label"] != df_eval["llm_label"]].copy()

    if df_errors.empty:
        logger.info("Không có lỗi phân loại — LLM khớp hoàn toàn với human_label!")
        df_errors = pd.DataFrame(columns=["source_id", "legal_citation",
                                          "human_label", "llm_label", "error_pair"])

    df_errors["error_pair"] = (
        df_errors["human_label"] + " → " + df_errors["llm_label"]
    )

    # Thống kê cặp lỗi
    error_counts = df_errors["error_pair"].value_counts().reset_index()
    error_counts.columns = ["Cặp lỗi (Human → LLM)", "Số record"]

    # Chọn cột hiển thị
    display_cols = [
        "source_id", "legal_citation", "raw_text",
        "human_label", "llm_label", "error_pair",
        "llm_reason", "llm_evidence_span", "llm_confidence",
        "human_reason",
    ]
    existing = [c for c in display_cols if c in df_errors.columns]
    df_errors_display = df_errors[existing]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_errors_display.to_excel(writer, index=False, sheet_name="Lỗi phân loại")
        error_counts.to_excel(writer, index=False, sheet_name="Thống kê cặp lỗi")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 22
            if "raw_text" in [c.value for c in ws[1]]:
                for i, cell in enumerate(ws[1], 1):
                    if cell.value == "raw_text":
                        from openpyxl.utils import get_column_letter
                        ws.column_dimensions[get_column_letter(i)].width = 80
                        break

    logger.info(f"Đã xuất error analysis ({len(df_errors)} lỗi) → {output_path}")
    return len(df_errors)


def build_final_labeled_dataset(df_merged: pd.DataFrame, mask: list) -> pd.DataFrame:
    """
    Tạo final_labeled_dataset:
      final_label = human_label (ưu tiên), fallback llm_label
      label_match = True nếu human_label == llm_label (chỉ với nhãn hợp lệ)
                    False nếu cả hai đều hợp lệ nhưng khác nhau
                    None nếu thiếu một trong hai nhãn
    """
    df = df_merged.copy()

    final_labels = []
    label_sources = []
    label_matches = []

    labels_set = set(LABELS_5)

    for _, row in df.iterrows():
        # Ép kiểu an toàn — tránh crash khi giá trị là float/NaN
        human_raw = row.get("human_label", None)
        llm_raw = row.get("llm_label", None)

        human = str(human_raw).strip().upper() if pd.notna(human_raw) else ""
        llm = str(llm_raw).strip().upper() if pd.notna(llm_raw) else ""

        human_valid = human in labels_set
        llm_valid = llm in labels_set

        if human_valid:
            final_labels.append(human)
            label_sources.append("human")
            if llm_valid:
                label_matches.append(human == llm)  # True hoặc False
            else:
                label_matches.append(None)  # Không có LLM hợp lệ để so sánh
        elif llm_valid:
            final_labels.append(llm)
            label_sources.append("llm")
            label_matches.append(None)  # Không có human_label để so sánh
        else:
            final_labels.append("UNKNOWN")
            label_sources.append("none")
            label_matches.append(None)

    df["final_label"] = final_labels
    df["final_label_source"] = label_sources
    df["label_match"] = label_matches
    return df


def main():
    parser = argparse.ArgumentParser(
        description="Bước 7: So sánh LLM vs Human — 8 metrics đầy đủ (5 nhãn)"
    )
    parser.add_argument("--config", type=Path,
                        default=PROJECT_ROOT / "config" / "project_config.yaml")
    parser.add_argument("--llm-file", type=Path, default=None)
    parser.add_argument("--human-file", type=Path, default=None)
    parser.add_argument("--output-metrics", type=Path, default=None)
    parser.add_argument("--output-errors", type=Path, default=None)
    parser.add_argument("--output-final", type=Path, default=None)
    args = parser.parse_args()

    logger.info("=" * 65)
    logger.info("  BƯỚC 7: SO SÁNH LLM vs HUMAN — 8 METRICS (5 NHÃN)")
    logger.info(f"  Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 65)

    config = load_config(args.config)
    interim_dir = PROJECT_ROOT / config["paths"]["interim_data"]
    processed_dir = PROJECT_ROOT / config["paths"]["processed_data"]
    reports_dir = PROJECT_ROOT / config["paths"]["reports"]

    llm_path = args.llm_file or (processed_dir / config["processed_files"]["llm_labeled_dataset"])
    human_path = args.human_file or (interim_dir / config["interim_files"]["human_labeled_dataset"])
    metrics_path = args.output_metrics or (reports_dir / config["report_files"]["classification_metrics"])
    errors_path = args.output_errors or (reports_dir / config["report_files"]["error_analysis"])
    final_path = args.output_final or (processed_dir / config["processed_files"]["final_labeled_dataset"])

    if not llm_path.exists():
        logger.error(f"Không tìm thấy file LLM: {llm_path}")
        logger.error("Chạy trước: python src/05_llm_classify_5_labels.py")
        sys.exit(1)

    df_llm = pd.read_excel(llm_path)
    logger.info(f"Đọc LLM dataset: {len(df_llm)} bản ghi")

    if not human_path.exists():
        logger.warning(f"Không tìm thấy file nhãn người: {human_path}")
        logger.warning("Chỉ xuất final_labeled_dataset, không tính metrics.")
        df_merged = df_llm.copy()
        df_merged["human_label"] = ""
    else:
        df_human = pd.read_excel(human_path)
        logger.info(f"Đọc human dataset: {len(df_human)} bản ghi")
        df_merged = merge_datasets(df_llm, df_human)

    # ── Tính metrics nếu có nhãn người ────────────────────────────────────────
    has_human = (
        "human_label" in df_merged.columns
        and df_merged["human_label"].isin(LABELS_5).any()
        and "llm_label" in df_merged.columns
    )

    mask_for_final = []
    if has_human:
        # Đảm bảo các cột nhãn là kiểu chuỗi trước khi dùng .str accessor
        # (tránh AttributeError khi cột chứa float/None xen lẫn string)
        y_true = (
            df_merged["human_label"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
            .values
        )
        y_pred = (
            df_merged["llm_label"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
            .values
        )

        metrics = compute_all_metrics(y_true, y_pred)
        if metrics:
            print_metrics_summary(metrics)
            export_metrics_excel(metrics, metrics_path)
            n_errors = export_error_analysis(df_merged, errors_path)
            logger.info(f"Số lỗi phân loại: {n_errors} / {metrics['M_eval']}")
            mask_for_final = metrics.get("mask", [])
    else:
        logger.warning("Chưa đủ nhãn người để tính metrics.")
        logger.warning("Điền human_label vào data/interim/human_labeled_dataset.xlsx rồi chạy lại.")

    # ── Tạo final_labeled_dataset ──────────────────────────────────────────────
    df_final = build_final_labeled_dataset(df_merged, mask_for_final)
    final_path.parent.mkdir(parents=True, exist_ok=True)
    df_final.to_excel(final_path, index=False, engine="openpyxl")
    logger.info(f"Đã xuất final_labeled_dataset → {final_path}")

    logger.info("")
    logger.info("✅ BƯỚC 7 HOÀN TẤT")
    logger.info("   BƯỚC TIẾP THEO:")
    logger.info("   python src/07_build_scoring_input.py")


if __name__ == "__main__":
    main()
