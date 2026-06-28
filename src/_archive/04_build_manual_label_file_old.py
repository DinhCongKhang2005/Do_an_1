#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
04_build_manual_label_file.py
==============================
BƯỚC 4 — Tạo file Excel để người nghiên cứu gán nhãn thủ công (5 nhãn).

Đề tài: Đánh giá tác động chính sách môi trường — Phiên bản 5 nhãn
Tác giả: Đinh Công Khang — MI3380 Đồ án 1

Pipeline:
  Input:  data/interim/environmental_impact_records.xlsx  (output của Bước 3)
  Output:
    - data/interim/manual_label_template.xlsx  (người nghiên cứu điền vào)

Sau khi chạy script này, người nghiên cứu cần:
  1. Mở data/interim/manual_label_template.xlsx
  2. Điền cột 'human_label' với MỘT trong 5 nhãn:
       BENEFIT_QUANTITATIVE
       BENEFIT_QUALITATIVE
       COST_QUANTITATIVE
       COST_QUALITATIVE
       CONSTRAINT
  3. Điền cột 'human_reason' — lý do ngắn gọn (1-2 câu)
  4. Điền cột 'review_note' — ghi chú nếu record khó/mơ hồ (có thể để trống)
  5. Đánh dấu 'needs_second_review' = TRUE nếu cần xem lại lần 2 (có thể để trống)
  6. Lưu file thành: data/interim/human_labeled_dataset.xlsx

Xem docs/label_guideline.md để biết quy tắc phân loại chi tiết.

Sử dụng:
  python src/04_build_manual_label_file.py
  python src/04_build_manual_label_file.py --input data/interim/environmental_impact_records.xlsx
  python src/04_build_manual_label_file.py --help
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml

PROJECT_ROOT = Path(__file__).resolve().parent.parent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

VALID_LABELS_5 = [
    "BENEFIT_QUANTITATIVE",
    "BENEFIT_QUALITATIVE",
    "COST_QUANTITATIVE",
    "COST_QUALITATIVE",
    "CONSTRAINT",
]


def load_config(config_path: Path) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def build_template(df: pd.DataFrame) -> pd.DataFrame:
    """
    Xây dựng file template gán nhãn thủ công 5 nhãn.
    Giữ các cột cần thiết cho context, thêm cột để người nghiên cứu điền.
    """
    # Chọn cột context cần thiết
    context_cols = [
        "source_id", "legal_citation", "raw_text",
        "chu_the", "tin_hieu_tac_dong", "domain",
        "gia_tri_dinh_luong", "dieu_kien_ap_dung",
        "muc_do_ro_rang", "ly_do",
    ]
    existing_context = [c for c in context_cols if c in df.columns]
    df_template = df[existing_context].copy()

    # Thêm cột gán nhãn (để trống — người nghiên cứu điền)
    df_template["human_label"] = ""
    df_template["human_reason"] = ""
    df_template["review_note"] = ""
    df_template["needs_second_review"] = ""

    return df_template


def export_template_excel(df: pd.DataFrame, output_path: Path):
    """Xuất template ra Excel với định dạng hỗ trợ gán nhãn."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="manual_label_template")
        ws = writer.sheets["manual_label_template"]

        # Định dạng độ rộng cột
        col_widths = {
            "A": 22,   # source_id
            "B": 22,   # legal_citation
            "C": 90,   # raw_text — rộng để đọc
            "D": 25,   # chu_the
            "E": 45,   # tin_hieu_tac_dong
            "F": 20,   # domain
            "G": 20,   # gia_tri_dinh_luong
            "H": 35,   # dieu_kien_ap_dung
            "I": 15,   # muc_do_ro_rang
            "J": 40,   # ly_do
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Tìm cột human_label và định dạng
        from openpyxl.utils import get_column_letter
        col_names = [cell.value for cell in ws[1]]

        for col_idx, col_name in enumerate(col_names, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name == "human_label":
                ws.column_dimensions[col_letter].width = 25
            elif col_name == "human_reason":
                ws.column_dimensions[col_letter].width = 55
            elif col_name == "review_note":
                ws.column_dimensions[col_letter].width = 40
            elif col_name == "needs_second_review":
                ws.column_dimensions[col_letter].width = 20

        # Đóng băng dòng tiêu đề
        ws.freeze_panes = "A2"

        # Thêm sheet hướng dẫn
        ws_guide = writer.book.create_sheet("Hướng dẫn gán nhãn")
        guide_content = [
            ["=== HƯỚNG DẪN GÁN NHÃN 5 LỚP ==="],
            [""],
            ["Nhãn", "Ý nghĩa", "Ví dụ"],
            ["BENEFIT_QUANTITATIVE", "Lợi ích có số, tiền, tỷ lệ % cụ thể", "Hỗ trợ 30% chi phí đầu tư..."],
            ["BENEFIT_QUALITATIVE", "Lợi ích tích cực nhưng chưa có số cụ thể", "Nhà nước khuyến khích tổ chức..."],
            ["COST_QUANTITATIVE", "Chi phí/phí có số tiền, tỷ lệ % cụ thể", "Phí bảo vệ môi trường ≥ 10% chi phí..."],
            ["COST_QUALITATIVE", "Nghĩa vụ thủ tục/tuân thủ chưa có số", "Chủ dự án phải lập báo cáo ĐTM"],
            ["CONSTRAINT", "Điều kiện, lệnh cấm, ngưỡng kỹ thuật, thời hạn", "Nghiêm cấm xả nước thải chưa xử lý"],
            [""],
            ["=== CẶP DỄ NHẦM ==="],
            ["COST_QUALITATIVE vs CONSTRAINT", "Nếu phải làm thủ tục → COST_QUALITATIVE. Nếu cấm đoán/ngưỡng/điều kiện → CONSTRAINT"],
            ["BENEFIT_QUANTITATIVE vs BENEFIT_QUALITATIVE", "Chỉ QUANTITATIVE nếu số gắn trực tiếp với lợi ích"],
            [""],
            ["=== CÁC CỘT CẦN ĐIỀN ==="],
            ["human_label", "BẮT BUỘC — Một trong 5 nhãn ở trên (viết hoa, đúng chính tả)"],
            ["human_reason", "Khuyến nghị — Lý do 1-2 câu"],
            ["review_note", "Tùy chọn — Ghi chú nếu record khó/mơ hồ"],
            ["needs_second_review", "Tùy chọn — Ghi TRUE nếu cần xem lại"],
            [""],
            ["Lưu file này thành: data/interim/human_labeled_dataset.xlsx"],
        ]
        for row_data in guide_content:
            ws_guide.append(row_data)

        ws_guide.column_dimensions["A"].width = 35
        ws_guide.column_dimensions["B"].width = 70
        ws_guide.column_dimensions["C"].width = 45

    logger.info(f"Đã xuất template → {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Bước 4: Tạo file template gán nhãn thủ công 5 nhãn"
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=PROJECT_ROOT / "config" / "project_config.yaml",
    )
    parser.add_argument(
        "--input", type=Path, default=None,
        help="File Excel đầu vào (mặc định: data/interim/environmental_impact_records.xlsx)"
    )
    parser.add_argument(
        "--output", type=Path, default=None,
        help="File Excel đầu ra (mặc định: data/interim/manual_label_template.xlsx)"
    )
    args = parser.parse_args()

    logger.info("=" * 65)
    logger.info("  BƯỚC 4: TẠO FILE GÁN NHÃN THỦ CÔNG (5 NHÃN)")
    logger.info(f"  Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 65)

    config = load_config(args.config)

    input_path = args.input or (
        PROJECT_ROOT / config["paths"]["interim_data"]
        / config["pipeline_outputs"]["environmental_impact_records"]
    )
    output_path = args.output or (
        PROJECT_ROOT / config["paths"]["interim_data"]
        / config["interim_files"]["manual_label_template"]
    )

    if not input_path.exists():
        logger.error(f"Không tìm thấy file đầu vào: {input_path}")
        logger.error("Chạy trước: python src/03_filter_environmental_impact.py")
        sys.exit(1)

    df = pd.read_excel(input_path)
    logger.info(f"Đọc được {len(df)} bản ghi từ: {input_path}")

    df_template = build_template(df)
    export_template_excel(df_template, output_path)

    logger.info("")
    logger.info("✅ BƯỚC 4 HOÀN TẤT")
    logger.info(f"   Số bản ghi cần gán nhãn: {len(df_template)}")
    logger.info(f"   File template: {output_path}")
    logger.info("")
    logger.info("   === BƯỚC THỦ CÔNG TIẾP THEO ===")
    logger.info("   1. Mở file: data/interim/manual_label_template.xlsx")
    logger.info("   2. Xem sheet 'Hướng dẫn gán nhãn' để biết quy tắc")
    logger.info("   3. Điền cột 'human_label' với 1 trong 5 nhãn:")
    for lbl in VALID_LABELS_5:
        logger.info(f"      - {lbl}")
    logger.info("   4. Điền 'human_reason', 'review_note', 'needs_second_review'")
    logger.info("   5. Lưu thành: data/interim/human_labeled_dataset.xlsx")
    logger.info("")
    logger.info("   Sau khi hoàn thành gán nhãn thủ công:")
    logger.info("   python src/05_llm_classify_5_labels.py")


if __name__ == "__main__":
    main()
