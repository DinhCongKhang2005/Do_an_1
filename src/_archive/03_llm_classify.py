#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
03_llm_classify.py
===================
BƯỚC 3 — Phân loại điều khoản pháp lý bằng một LLM duy nhất.

Đề tài: Đánh giá tác động chính sách môi trường — Phiên bản không tranh biện
Tác giả: Đinh Công Khang — MI3380 Đồ án 1

Pipeline:
  Input:  data/interim/impact_true_records.xlsx (hoặc human_labeled_dataset.xlsx nếu có)
  Output: data/processed/llm_labeled_dataset.xlsx

Yêu cầu:
  - File .env với GOOGLE_API_KEY (hoặc GOOGLE_CLOUD_PROJECT cho Vertex AI)
  - Prompt tại: prompts/classify_system_prompt.txt

Sử dụng:
  python src/03_llm_classify.py
  python src/03_llm_classify.py --input data/interim/human_labeled_dataset.xlsx
  python src/03_llm_classify.py --dry-run  (kiểm tra cấu hình không gọi API)
  python src/03_llm_classify.py --help

Lưu ý:
  - Script này KHÔNG dùng Multi-Agent Debate
  - Kết quả LLM cần được nhà nghiên cứu xem xét trước khi dùng
"""

import argparse
import json
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml

# ─── Đường dẫn gốc dự án ─────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent

# ─── Cấu hình logging ─────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# ─── Nạp biến môi trường ─────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / ".env")
except ImportError:
    pass

VALID_LABELS = {"BENEFIT", "COST", "CONSTRAINT"}


def load_config(config_path: Path) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_system_prompt(prompt_path: Path) -> str:
    """Đọc file system prompt cho LLM."""
    if not prompt_path.exists():
        logger.warning(f"Không tìm thấy prompt tại: {prompt_path}")
        logger.warning("Dùng prompt dự phòng đơn giản.")
        return (
            "Bạn là chuyên gia phân loại điều khoản pháp luật môi trường. "
            "Phân loại thành đúng một trong ba nhãn: BENEFIT, COST, CONSTRAINT. "
            "BENEFIT: tác động tích cực. COST: chi phí/gánh nặng. "
            "CONSTRAINT: ngưỡng/giới hạn kỹ thuật. "
            "Ưu tiên CONSTRAINT nếu có giá trị số biểu thị ngưỡng kỹ thuật. "
            "Trả về JSON hợp lệ: {\"label\": \"BENEFIT|COST|CONSTRAINT\", "
            "\"reason\": \"...\", \"evidence_span\": \"...\", \"confidence\": 0.0}"
        )
    with open(prompt_path, "r", encoding="utf-8") as f:
        return f.read().strip()


def init_llm_client(provider: str):
    """
    Khởi tạo LLM client dựa trên provider.
    Hỗ trợ: 'google_ai_studio' hoặc 'vertex_ai'.
    """
    if provider == "google_ai_studio":
        try:
            import google.generativeai as genai
            api_key = os.environ.get("GOOGLE_API_KEY")
            if not api_key:
                raise ValueError("Thiếu GOOGLE_API_KEY trong file .env")
            genai.configure(api_key=api_key)
            return ("google_ai_studio", genai)
        except ImportError:
            raise ImportError("Thiếu thư viện: pip install google-generativeai")

    elif provider == "vertex_ai":
        try:
            from google import genai as vertex_genai
            project = os.environ.get("GOOGLE_CLOUD_PROJECT")
            location = os.environ.get("GOOGLE_CLOUD_LOCATION", "us-central1")
            if not project:
                raise ValueError("Thiếu GOOGLE_CLOUD_PROJECT trong file .env")
            client = vertex_genai.Client(vertexai=True, project=project, location=location)
            return ("vertex_ai", client)
        except ImportError:
            raise ImportError("Thiếu thư viện: pip install google-cloud-aiplatform")

    else:
        raise ValueError(f"Provider không hợp lệ: {provider}. Dùng 'google_ai_studio' hoặc 'vertex_ai'")


def call_llm_single(
    text: str,
    system_prompt: str,
    client_info: tuple,
    model_name: str,
    temperature: float = 0.0,
) -> dict:
    """
    Gọi một LLM duy nhất để phân loại điều khoản.
    Trả về dict với: label, reason, evidence_span, confidence.

    Lưu ý: Đây là phân loại đơn lẻ — KHÔNG có Multi-Agent Debate.
    """
    provider, client = client_info

    # Xây dựng user prompt
    user_prompt = (
        f"Hãy phân loại điều khoản pháp lý sau:\n\n"
        f"{text}\n\n"
        f"Trả về JSON hợp lệ theo yêu cầu trong system prompt."
    )

    raw_content = ""
    if provider == "google_ai_studio":
        import google.generativeai as genai
        model = genai.GenerativeModel(
            model_name=model_name,
            system_instruction=system_prompt,
            generation_config=genai.GenerationConfig(
                temperature=temperature,
                response_mime_type="application/json",
            ),
        )
        response = model.generate_content(user_prompt)
        raw_content = response.text.strip()

    elif provider == "vertex_ai":
        from google.genai import types
        response = client.models.generate_content(
            model=model_name,
            contents=user_prompt,
            config=types.GenerateContentConfig(
                system_instruction=system_prompt,
                temperature=temperature,
                response_mime_type="application/json",
            ),
        )
        raw_content = response.text.strip()

    # Làm sạch markdown nếu có
    if raw_content.startswith("```json"):
        raw_content = raw_content[7:]
    if raw_content.startswith("```"):
        raw_content = raw_content[3:]
    if raw_content.endswith("```"):
        raw_content = raw_content[:-3]

    result = json.loads(raw_content.strip())
    return result


def normalize_llm_label(raw_label: str) -> str:
    """
    Chuẩn hóa nhãn trả về từ LLM về một trong ba nhãn hợp lệ.
    """
    if not raw_label:
        return "UNKNOWN"
    upper = raw_label.strip().upper()
    # Khớp chính xác
    if upper in VALID_LABELS:
        return upper
    # Khớp từng phần
    for valid in VALID_LABELS:
        if valid in upper or upper in valid:
            return valid
    # Fallback theo tiếng Việt
    lower = raw_label.strip().lower()
    if any(k in lower for k in ["lợi", "benefit", "ích", "tích cực"]):
        return "BENEFIT"
    if any(k in lower for k in ["chi phí", "cost", "gánh nặng", "phí"]):
        return "COST"
    if any(k in lower for k in ["ràng buộc", "constraint", "ngưỡng", "giới hạn"]):
        return "CONSTRAINT"
    return "UNKNOWN"


def classify_dataframe(
    df: pd.DataFrame,
    system_prompt: str,
    client_info: tuple,
    model_name: str,
    text_field: str,
    delay_seconds: float = 0.5,
    dry_run: bool = False,
) -> pd.DataFrame:
    """
    Phân loại tất cả bản ghi trong DataFrame bằng LLM đơn.
    """
    total = len(df)
    logger.info(f"Bắt đầu phân loại {total} bản ghi với model: {model_name}")
    logger.info(f"(Chế độ: {'DRY RUN — không gọi API' if dry_run else 'THỰC TẾ — gọi API'})")

    llm_labels = []
    llm_reasons = []
    llm_evidences = []
    llm_confidences = []
    llm_errors = []

    for idx, row in df.iterrows():
        source_id = str(row.get("source_id", f"idx_{idx}"))
        text = str(row.get(text_field, "")).strip()

        print(f"  [{idx+1:>4}/{total}] {source_id[:25]:<25} ", end="", flush=True)

        if dry_run:
            # Chế độ kiểm tra không gọi API
            llm_labels.append("BENEFIT")
            llm_reasons.append("[DRY RUN] Không gọi API")
            llm_evidences.append("")
            llm_confidences.append(0.0)
            llm_errors.append("")
            print("→ [DRY RUN]")
            continue

        if not text:
            llm_labels.append("UNKNOWN")
            llm_reasons.append("Văn bản trống")
            llm_evidences.append("")
            llm_confidences.append(0.0)
            llm_errors.append("empty_text")
            print("→ ⚠️ Văn bản trống")
            continue

        try:
            result = call_llm_single(
                text=text,
                system_prompt=system_prompt,
                client_info=client_info,
                model_name=model_name,
            )

            raw_label = result.get("label", "UNKNOWN")
            normalized = normalize_llm_label(raw_label)

            llm_labels.append(normalized)
            llm_reasons.append(result.get("reason", ""))
            llm_evidences.append(result.get("evidence_span", ""))
            llm_confidences.append(float(result.get("confidence", 0.5)))
            llm_errors.append("")

            confidence = float(result.get("confidence", 0.5))
            flag = "⚠️" if confidence < 0.7 else "✅"
            print(f"→ {flag} {normalized} (conf={confidence:.2f})")

        except json.JSONDecodeError as e:
            error_msg = f"JSON parse error: {str(e)[:80]}"
            logger.error(f"Lỗi JSON dòng {idx}: {error_msg}")
            llm_labels.append("PARSE_ERROR")
            llm_reasons.append(error_msg)
            llm_evidences.append("")
            llm_confidences.append(0.0)
            llm_errors.append("json_parse_error")
            print(f"→ ❌ JSON Error")

        except Exception as e:
            error_msg = str(e)[:120]
            logger.error(f"Lỗi API dòng {idx}: {error_msg}")
            llm_labels.append("API_ERROR")
            llm_reasons.append(f"Lỗi API: {error_msg}")
            llm_evidences.append("")
            llm_confidences.append(0.0)
            llm_errors.append("api_error")
            print(f"→ ❌ API Error: {error_msg[:40]}")

        time.sleep(delay_seconds)

    df_result = df.copy()
    df_result["llm_label"] = llm_labels
    df_result["llm_reason"] = llm_reasons
    df_result["llm_evidence_span"] = llm_evidences
    df_result["llm_confidence"] = llm_confidences
    df_result["llm_error"] = llm_errors

    return df_result


def print_summary(df: pd.DataFrame):
    """In thống kê phân phối nhãn LLM."""
    logger.info(f"\n{'='*50}")
    logger.info(f"  THỐNG KÊ NHÃN LLM")
    logger.info(f"{'='*50}")
    counts = df["llm_label"].value_counts()
    for label, count in counts.items():
        bar = "█" * min(count, 30)
        logger.info(f"  {label:<15}: {count:>4}  {bar}")
    errors = df[df["llm_error"] != ""]["llm_error"].count()
    if errors > 0:
        logger.info(f"  {'LỖI':<15}: {errors:>4}  *** Cần kiểm tra lại ***")
    logger.info(f"{'='*50}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Bước 3: Phân loại điều khoản pháp lý bằng LLM đơn (không debate)"
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=PROJECT_ROOT / "config" / "project_config.yaml",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="File Excel đầu vào (mặc định: data/interim/impact_true_records.xlsx)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="File Excel đầu ra (mặc định: data/processed/llm_labeled_dataset.xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Chạy thử không gọi API (kiểm tra cấu hình)",
    )
    args = parser.parse_args()

    # ── Đọc cấu hình ──────────────────────────────────────────────────────────
    logger.info(f"{'='*65}")
    logger.info(f"  BƯỚC 3: PHÂN LOẠI BẰNG LLM ĐƠN (KHÔNG DEBATE)")
    logger.info(f"  Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"{'='*65}")

    config = load_config(args.config)
    interim_dir = PROJECT_ROOT / config["paths"]["interim_data"]
    processed_dir = PROJECT_ROOT / config["paths"]["processed_data"]

    # Đường dẫn file
    # Ưu tiên dùng human_labeled_dataset nếu có (để khớp source_id khi so sánh)
    human_file = interim_dir / config["interim_files"]["human_labeled_dataset"]
    impact_file = interim_dir / config["interim_files"]["impact_true_records"]

    if args.input:
        input_path = args.input
    elif human_file.exists():
        input_path = human_file
        logger.info(f"Phát hiện file đã gán nhãn thủ công: {input_path}")
    else:
        input_path = impact_file
        logger.info(f"Dùng file lọc (chưa có nhãn người): {input_path}")

    output_path = args.output or (processed_dir / config["processed_files"]["llm_labeled_dataset"])

    # Cấu hình LLM
    model_name = os.environ.get("LLM_MODEL_NAME", config["llm"].get("model", "gemini-2.5-flash"))
    temperature = float(os.environ.get("LLM_TEMPERATURE", config["llm"].get("temperature", 0.0)))
    delay_seconds = float(os.environ.get("LLM_DELAY_SECONDS", config["llm"].get("delay_seconds", 0.5)))
    text_field = config["llm"].get("text_field", "noi_dung_dieu_khoan")
    provider = os.environ.get("LLM_PROVIDER", "google_ai_studio")

    prompt_path = PROJECT_ROOT / config["paths"]["prompts"] / "classify_system_prompt.txt"

    # ── Kiểm tra file đầu vào ─────────────────────────────────────────────────
    if not input_path.exists():
        logger.error(f"Không tìm thấy file đầu vào: {input_path}")
        logger.error("Chạy trước: python src/01_load_and_filter_impact_true.py")
        sys.exit(1)

    # ── Đọc dữ liệu ───────────────────────────────────────────────────────────
    df = pd.read_excel(input_path)
    logger.info(f"Đọc được {len(df)} bản ghi từ: {input_path}")

    # ── Nạp prompt ────────────────────────────────────────────────────────────
    system_prompt = load_system_prompt(prompt_path)
    logger.info(f"Đã nạp system prompt ({len(system_prompt)} ký tự)")

    # ── Khởi tạo LLM client ───────────────────────────────────────────────────
    client_info = None
    if not args.dry_run:
        try:
            client_info = init_llm_client(provider)
            logger.info(f"Đã khởi tạo LLM client (provider={provider}, model={model_name})")
        except Exception as e:
            logger.error(f"Lỗi khởi tạo LLM: {e}")
            logger.error("Kiểm tra file .env và thư viện đã cài.")
            sys.exit(1)
    else:
        client_info = ("dry_run", None)

    # ── Phân loại ─────────────────────────────────────────────────────────────
    df_result = classify_dataframe(
        df=df,
        system_prompt=system_prompt,
        client_info=client_info,
        model_name=model_name,
        text_field=text_field,
        delay_seconds=delay_seconds,
        dry_run=args.dry_run,
    )

    # ── Xuất kết quả ─────────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_result.to_excel(writer, index=False, sheet_name="llm_labeled")
        ws = writer.sheets["llm_labeled"]
        ws.column_dimensions["A"].width = 30   # source_id
        # Tìm và mở rộng cột nội dung
        for col_idx, col_name in enumerate(df_result.columns, start=1):
            if col_name == "noi_dung_dieu_khoan":
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = 80
            elif col_name in ("llm_label",):
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            elif col_name in ("llm_reason", "llm_evidence_span"):
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = 50

    print_summary(df_result)

    logger.info(f"✅ BƯỚC 3 HOÀN TẤT")
    logger.info(f"   Kết quả LLM: {output_path}")
    logger.info(f"   Bản ghi phân loại: {len(df_result)}")
    logger.info(f"")
    logger.info(f"   BƯỚC TIẾP THEO:")
    logger.info(f"   python src/04_compare_llm_vs_human.py")


if __name__ == "__main__":
    main()
