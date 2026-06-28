#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
05_build_scoring_input.py
==========================
BƯỚC 5 — Tạo bảng biến đầu vào cho tính điểm Impact Score.

Đề tài: Đánh giá tác động chính sách môi trường — Phiên bản không tranh biện
Tác giả: Đinh Công Khang — MI3380 Đồ án 1

Pipeline:
  Input:  data/processed/final_labeled_dataset.xlsx
  Output: data/interim/scoring_input.xlsx  (chờ nhà nghiên cứu điền M/S/D/R)

Sau khi chạy:
  → Nhà nghiên cứu điền M_i, S_i, D_i, R_i (thang 1-5) và object_weight
  → Lưu file (không đổi tên)
  → Chạy: python src/06_calculate_impact_score.py

Sử dụng:
  python src/05_build_scoring_input.py
  python src/05_build_scoring_input.py --help
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml

PROJECT_ROOT = Path(__file__).resolve().parent.parent

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# Hướng theo nhãn
DIRECTION_MAP = {
    "BENEFIT": 1,
    "COST": -1,
    "CONSTRAINT": -1,
}

SCORING_GUIDE = """HƯỚNG DẪN ĐIỀN BIẾN TÍNH ĐIỂM IMPACT SCORE
=============================================
Điền các giá trị 1–5 vào các cột biến (thang Likert):

M_i  — Magnitude (Độ lớn tác động)
  1=Rất thấp | 2=Thấp | 3=Trung bình | 4=Cao | 5=Rất cao

S_i  — Scope (Phạm vi địa lý/đối tượng)
  1=Cá nhân | 2=Địa phương | 3=Vùng | 4=Quốc gia | 5=Toàn cầu

D_i  — Duration (Thời gian tác động)
  1=Tức thời | 2=Ngắn hạn (<1 năm) | 3=Trung hạn (1-5 năm) | 4=Dài hạn (>5 năm) | 5=Vĩnh viễn

R_i  — Risk/Reversibility (Khó khôi phục)
  1=Dễ khôi phục hoàn toàn | 2=Có thể khôi phục | 3=Trung bình | 4=Khó khôi phục | 5=Không thể khôi phục

object_weight — Trọng số đối tượng chịu tác động
  Mặc định: 1.0 (có thể điều chỉnh theo tầm quan trọng của đối tượng)

Sau khi điền xong, lưu file scoring_input.xlsx và chạy:
  python src/06_calculate_impact_score.py
"""


def load_config(config_path: Path) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def build_scoring_input(df: pd.DataFrame, default_object_weight: float = 1.0) -> pd.DataFrame:
    """
    Tạo bảng biến tính điểm từ final_labeled_dataset.
    Chỉ bao gồm các bản ghi có final_label hợp lệ.
    Thêm các cột biến cần nhà nghiên cứu điền.
    """
    valid_labels = {"BENEFIT", "COST", "CONSTRAINT"}

    # Lọc bản ghi có nhãn hợp lệ
    mask = df["final_label"].str.strip().str.upper().isin(valid_labels)
    df_valid = df[mask].copy()

    logger.info(f"Bản ghi có final_label hợp lệ: {len(df_valid)}/{len(df)}")

    if df_valid.empty:
        logger.warning("Không có bản ghi nào có nhãn hợp lệ để tính điểm.")
        return df_valid

    # Cột cơ bản cần giữ
    keep_cols = [
        "source_id", "legal_citation", "noi_dung_dieu_khoan",
        "final_label", "final_label_source",
        "llm_label", "llm_reason", "llm_confidence",
        "human_label", "human_reason",
    ]
    existing_keep = [c for c in keep_cols if c in df_valid.columns]
    df_out = df_valid[existing_keep].copy()

    # Thêm cột hướng tác động (tự tính từ final_label)
    df_out["direction"] = df_out["final_label"].str.strip().str.upper().map(DIRECTION_MAP).fillna(0).astype(int)

    # Thêm các cột biến Likert để nhà nghiên cứu điền
    df_out.insert(df_out.columns.get_loc("direction") + 1, "M_i", "")  # Magnitude
    df_out.insert(df_out.columns.get_loc("M_i") + 1, "S_i", "")        # Scope
    df_out.insert(df_out.columns.get_loc("S_i") + 1, "D_i", "")        # Duration
    df_out.insert(df_out.columns.get_loc("D_i") + 1, "R_i", "")        # Risk/Reversibility
    df_out.insert(df_out.columns.get_loc("R_i") + 1, "object_weight", default_object_weight)
    df_out.insert(df_out.columns.get_loc("object_weight") + 1, "affected_object", "")  # Tên đối tượng

    return df_out


def export_scoring_input_excel(df: pd.DataFrame, output_path: Path):
    """Xuất bảng biến tính điểm ra Excel với hướng dẫn."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Biến tính điểm")
        ws = writer.sheets["Biến tính điểm"]

        # Điều chỉnh độ rộng cột
        ws.column_dimensions["A"].width = 30   # source_id
        ws.column_dimensions["B"].width = 22   # legal_citation
        ws.column_dimensions["C"].width = 80   # noi_dung_dieu_khoan
        ws.column_dimensions["D"].width = 14   # final_label
        ws.column_dimensions["E"].width = 10   # final_label_source
        ws.column_dimensions["F"].width = 12   # direction

        # Cột biến (M/S/D/R)
        for col_name in ["M_i", "S_i", "D_i", "R_i"]:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = 8

        # Thêm sheet hướng dẫn
        ws_guide = writer.book.create_sheet(title="Hướng dẫn")
        for row_idx, line in enumerate(SCORING_GUIDE.strip().split("\n"), start=1):
            ws_guide.cell(row=row_idx, column=1, value=line)
        ws_guide.column_dimensions["A"].width = 80

    logger.info(f"Đã xuất bảng biến tính điểm → {output_path} ({len(df)} bản ghi)")


def main():
    parser = argparse.ArgumentParser(
        description="Bước 5: Tạo bảng biến đầu vào tính điểm Impact Score"
    )
    parser.add_argument("--config", type=Path, default=PROJECT_ROOT / "config" / "project_config.yaml")
    parser.add_argument("--input", type=Path, default=None, help="File final_labeled_dataset.xlsx")
    parser.add_argument("--output", type=Path, default=None, help="File scoring_input.xlsx đầu ra")
    args = parser.parse_args()

    logger.info(f"{'='*65}")
    logger.info(f"  BƯỚC 5: TẠO BẢNG BIẾN TÍNH ĐIỂM IMPACT SCORE")
    logger.info(f"  Thời gian: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"{'='*65}")

    config = load_config(args.config)
    processed_dir = PROJECT_ROOT / config["paths"]["processed_data"]
    interim_dir = PROJECT_ROOT / config["paths"]["interim_data"]

    input_path = args.input or (processed_dir / config["processed_files"]["final_labeled_dataset"])
    output_path = args.output or (interim_dir / config["interim_files"]["scoring_input"])
    default_weight = config["scoring"].get("default_object_weight", 1.0)

    if not input_path.exists():
        logger.error(f"Không tìm thấy file: {input_path}")
        logger.error("Chạy trước: python src/04_compare_llm_vs_human.py")
        sys.exit(1)

    df = pd.read_excel(input_path)
    logger.info(f"Đọc được {len(df)} bản ghi từ: {input_path}")

    df_scoring = build_scoring_input(df, default_object_weight=default_weight)

    if df_scoring.empty:
        logger.error("Không có dữ liệu để xuất. Kiểm tra cột 'final_label' trong input file.")
        sys.exit(1)

    export_scoring_input_excel(df_scoring, output_path)

    logger.info(f"")
    logger.info(f"✅ BƯỚC 5 HOÀN TẤT")
    logger.info(f"   Bảng biến đã tạo: {output_path} ({len(df_scoring)} điều khoản)")
    logger.info(f"")
    logger.info(f"   *** DỪNG — BƯỚC THỦ CÔNG ***")
    logger.info(f"   1. Mở file: {output_path}")
    logger.info(f"   2. Điền M_i, S_i, D_i, R_i (thang 1-5) và object_weight cho từng điều khoản")
    logger.info(f"   3. Lưu file (không đổi tên)")
    logger.info(f"   4. Chạy: python src/06_calculate_impact_score.py")


if __name__ == "__main__":
    main()
