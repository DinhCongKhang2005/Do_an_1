#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
06_build_class_manual_label_file.py
=====================================
BƯỚC 6 — Tạo file template để người nghiên cứu gán nhãn class_human (5 lớp).

Đề tài: Đánh giá tác động chính sách môi trường — Pipeline 2 tầng
Tác giả: Đinh Công Khang — MI3380 Đồ án 1

Vị trí trong pipeline:
    Input:  data/processed/env_final_dataset.xlsx  (D_env, chỉ lấy env_final=1)
    Output: data/interim/class_manual_label_template.xlsx
    Tiếp theo: [HUMAN gán class_human] → Script 07

Lý do cần bước này:
    - Tầng 2 chỉ xử lý D_env = {bản ghi có env_final=1}
    - Template cần cột trống để gán 1 trong 5 nhãn:
      BENEFIT_QUANTITATIVE | BENEFIT_QUALITATIVE |
      COST_QUANTITATIVE | COST_QUALITATIVE | CONSTRAINT
    - Hiển thị kết quả Tầng 1 (env_final, env_llm_reason) để tham khảo

Hướng dẫn sau khi chạy:
    1. Đọc docs/label_guideline_5_labels.md
    2. Đọc docs/adjudication_protocol.md (quy trình)
    3. Mở class_manual_label_template.xlsx
    4. Điền cột class_human: 1 trong 5 nhãn (viết đúng chính xác tên nhãn)
    5. Điền class_human_reason: lý do + quy tắc dùng
    6. Điền class_needs_review: True nếu còn mơ hồ
    7. Lưu thành data/interim/class_human_labeled_dataset.xlsx

Sử dụng:
    py src/06_build_class_manual_label_file.py
"""

import logging
import sys
from pathlib import Path

import pandas as pd
import yaml

# ─── Project root ──────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        pass

VALID_5_LABELS = [
    "BENEFIT_QUANTITATIVE",
    "BENEFIT_QUALITATIVE",
    "COST_QUANTITATIVE",
    "COST_QUALITATIVE",
    "CONSTRAINT",
]


def load_config():
    with open(PROJECT_ROOT / "config" / "project_config.yaml", encoding="utf-8") as f:
        return yaml.safe_load(f)


def main():
    cfg = load_config()
    
    logger.info("=" * 60)
    logger.info("  BƯỚC 6 — Tạo template gán nhãn class_human (5 lớp)")
    logger.info("=" * 60)
    
    processed_dir = PROJECT_ROOT / cfg["paths"]["processed_data"]
    interim_dir   = PROJECT_ROOT / cfg["paths"]["interim_data"]
    
    # ── Đọc D_env ──
    env_path = processed_dir / cfg["processed_files"]["env_final_dataset"]
    if not env_path.exists():
        logger.error(f"Không tìm thấy: {env_path}")
        logger.error("Chạy script 05 trước!")
        sys.exit(1)
    
    logger.info(f"[Bước 6.1] Đọc D_env: {env_path.name}")
    df_env = pd.read_excel(env_path, dtype=str)
    
    # Chỉ lấy bản ghi env_final = 1
    df = df_env[df_env["env_final"].astype(str).str.strip().isin(["1", "True", "true"])].copy()
    logger.info(f"  → D_env (env_final=1): {len(df)} bản ghi")
    
    # ── Chọn cột hiển thị ──
    base_cols = ["source_id", "legal_citation", "raw_text", "actor",
                 "legal_signal", "domain", "quantitative_value", "condition"]
    env_info_cols = [c for c in ["env_final", "env_llm", "env_llm_reason", "env_evidence_span",
                                  "env_confidence", "env_llm_domain"] if c in df.columns]
    
    df_template = df[base_cols + env_info_cols].copy()
    
    # ── Thêm cột cần điền ──
    df_template["class_human"]        = ""    # Một trong 5 nhãn
    df_template["class_human_reason"] = ""    # Lý do + quy tắc đã dùng
    df_template["class_needs_review"] = ""    # True/False
    
    # ── Thêm cột nhắc nhở ──
    df_template["VALID_LABELS"] = " | ".join(VALID_5_LABELS)  # Reminder nhãn hợp lệ
    
    # ── Lưu file ──
    output_path = interim_dir / cfg["interim_files"]["class_manual_label_template"]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    logger.info(f"[Bước 6.2] Lưu template: {output_path.name}")
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_template.to_excel(writer, sheet_name="class_label_template", index=False)
        
        # Định dạng
        ws = writer.sheets["class_label_template"]
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Độ rộng cột
        col_config = {
            1: 20,   # source_id
            2: 30,   # legal_citation
            3: 65,   # raw_text
            4: 25,   # actor
            5: 20,   # legal_signal
            6: 20,   # domain
            7: 15,   # quantitative_value
            8: 25,   # condition
        }
        import openpyxl.utils as utils_xl
        for col_idx, width in col_config.items():
            col_letter = utils_xl.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
        
        # Wrap text raw_text
        for row in ws.iter_rows(min_row=2):
            row[2].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Highlight cột cần điền (class_human, class_human_reason, class_needs_review)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        n_base = len(base_cols) + len(env_info_cols)
        for col_offset in [1, 2, 3]:  # class_human, class_human_reason, class_needs_review
            cell = ws.cell(row=1, column=n_base + col_offset)
            cell.fill = yellow_fill
            cell.font = bold_font
        
        # Thêm sheet nhắc nhở labels
        label_df = pd.DataFrame({
            "Nhãn": VALID_5_LABELS,
            "Mã ngắn": ["BQ", "BQL", "CQ", "CQL", "CON"],
            "Mô tả": [
                "Lợi ích định lượng — có tỉ lệ %, khối lượng, nồng độ cụ thể liên quan MT",
                "Lợi ích định tính — có lợi ích MT nhưng không có ngưỡng số cụ thể",
                "Chi phí định lượng — có nghĩa vụ tuân thủ + ngưỡng số môi trường",
                "Chi phí định tính — có nghĩa vụ thủ tục (lập/trình/xin phép) nhưng không có ngưỡng số",
                "Ràng buộc — cấm/bắt buộc hành vi, không tạo lợi ích rõ ràng, không đo chi phí",
            ],
            "direction": ["+1", "+1", "-1", "-1", "-1"]
        })
        label_df.to_excel(writer, sheet_name="labels_reference", index=False)
    
    logger.info(f"  ✅ Đã lưu: {output_path.name} ({len(df_template)} dòng)")
    logger.info("")
    logger.info("=" * 60)
    logger.info("  HOÀN THÀNH — Hướng dẫn tiếp theo:")
    logger.info(f"  1. Đọc docs/label_guideline_5_labels.md")
    logger.info(f"  2. Mở {output_path.name}")
    logger.info(f"  3. Điền class_human (viết ĐÚNG tên nhãn, không viết tắt)")
    logger.info(f"  4. Lưu thành: class_human_labeled_dataset.xlsx")
    logger.info(f"  5. Chạy: py src/07_llm_classify_5_labels.py")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
